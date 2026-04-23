"""
Customer API: Excel (.xlsx) by default, or CSV if CUSTOMER_DATA_PATH ends with .csv.
WireMock reverse-proxies /customers* here (see wiremock/mappings).

Set CUSTOMER_DATA_PATH to override the file path.
"""
from __future__ import annotations

import csv
import logging
import os
import threading
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from logging.handlers import RotatingFileHandler

from flask import Flask, Response, g, jsonify, request
from openpyxl import Workbook, load_workbook

APP_DIR = Path(__file__).resolve().parent
_default_xlsx = APP_DIR / "data" / "customers.xlsx"
DATA_PATH = Path(os.environ.get("CUSTOMER_DATA_PATH", _default_xlsx))

FIELDNAMES = [
    "customer_id",
    "first_name",
    "last_name",
    "email",
    "phone",
    "status",
    "registered_at",
]

_file_lock = threading.Lock()

LOG_DIR = APP_DIR / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_PATH = LOG_DIR / "backend.log"

logger = logging.getLogger("customer-backend")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter(
        fmt="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%dT%H:%M:%SZ",
    )
    file_handler = RotatingFileHandler(
        LOG_PATH, maxBytes=2_000_000, backupCount=3, encoding="utf-8"
    )
    file_handler.setFormatter(fmt)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(fmt)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _is_xlsx() -> bool:
    return DATA_PATH.suffix.lower() in (".xlsx", ".xlsm")


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        dt = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    if isinstance(value, date):
        return value.isoformat() + "T00:00:00Z"
    return str(value).strip()


def _seed_rows() -> list[dict[str, str]]:
    now = _utc_now_iso()
    return [
        {
            "customer_id": "cust-001",
            "first_name": "John",
            "last_name": "Doe",
            "email": "john.doe@example.com",
            "phone": "+1-202-555-0123",
            "status": "ACTIVE",
            "registered_at": now,
        },
        {
            "customer_id": "cust-002",
            "first_name": "Jane",
            "last_name": "Roe",
            "email": "jane.roe@example.com",
            "phone": "+1-202-555-0456",
            "status": "ACTIVE",
            "registered_at": now,
        },
        {
            "customer_id": "cust-003",
            "first_name": "Sam",
            "last_name": "Lee",
            "email": "sam.lee@example.com",
            "phone": "",
            "status": "INACTIVE",
            "registered_at": now,
        },
    ]


def read_rows_csv() -> list[dict[str, str]]:
    if not DATA_PATH.exists():
        return []
    with DATA_PATH.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return []
        return [
            {k: (row.get(k) or "").strip() for k in FIELDNAMES} for row in reader
        ]


def write_rows_csv(rows: list[dict[str, str]]) -> None:
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    with DATA_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in FIELDNAMES})


def read_rows_xlsx() -> list[dict[str, str]]:
    if not DATA_PATH.exists():
        return []
    wb = load_workbook(DATA_PATH, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_out: list[dict[str, str]] = []
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row or not any(cell is not None and str(cell).strip() for cell in row):
                continue
            rec: dict[str, str] = {}
            for j, key in enumerate(FIELDNAMES):
                val = row[j] if j < len(row) else None
                rec[key] = _cell_str(val)
            rows_out.append(rec)
        return rows_out
    finally:
        wb.close()


def write_rows_xlsx(rows: list[dict[str, str]]) -> None:
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(list(FIELDNAMES))
    for r in rows:
        ws.append([r.get(k, "") for k in FIELDNAMES])
    wb.save(DATA_PATH)


def read_rows() -> list[dict[str, str]]:
    if _is_xlsx():
        return read_rows_xlsx()
    return read_rows_csv()


def write_rows(rows: list[dict[str, str]]) -> None:
    if _is_xlsx():
        write_rows_xlsx(rows)
    else:
        write_rows_csv(rows)


def bootstrap_data() -> None:
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DATA_PATH.exists() and DATA_PATH.stat().st_size > 0:
        return
    write_rows(_seed_rows())


def row_to_json(row: dict[str, str]) -> dict:
    phone = row.get("phone", "").strip()
    return {
        "customerId": row["customer_id"],
        "firstName": row["first_name"],
        "lastName": row["last_name"],
        "email": row["email"],
        "phone": phone or None,
        "status": row["status"],
        "registeredAt": row["registered_at"],
    }


def create_app() -> Flask:
    app = Flask(__name__)

    @app.before_request
    def before_request() -> None:
        rid = request.headers.get("X-Request-Id") or str(uuid.uuid4())
        g.request_id = rid
        logger.info(
            "IN %s %s from=%s requestId=%s",
            request.method,
            request.path,
            request.remote_addr,
            rid,
        )

    @app.after_request
    def after_request(resp: Response) -> Response:
        rid = getattr(g, "request_id", None)
        if rid:
            resp.headers["X-Request-Id"] = str(rid)
            logger.info(
                "OUT %s %s status=%s requestId=%s",
                request.method,
                request.path,
                resp.status_code,
                rid,
            )
        return resp

    @app.route("/health", methods=["GET"])
    def health() -> Response:
        return jsonify({"status": "ok"})

    @app.route("/customers/<customer_id>", methods=["GET"])
    def get_customer(customer_id: str) -> tuple[Response, int] | Response:
        with _file_lock:
            rows = read_rows()
        for row in rows:
            if row["customer_id"] == customer_id:
                logger.info("GET hit customer_id=%s email=%s", customer_id, row.get("email", ""))
                return jsonify(row_to_json(row))
        logger.info("GET miss customer_id=%s", customer_id)
        return jsonify({"error": "Customer not found"}), 404

    @app.route("/customers", methods=["POST"])
    def create_customer() -> tuple[Response, int] | Response:
        if not request.is_json:
            return jsonify({"error": "Content-Type must be application/json"}), 400
        body = request.get_json(silent=True) or {}
        first = body.get("firstName")
        last = body.get("lastName")
        email_raw = (body.get("email") or "").strip()
        if not first or not last or not email_raw:
            return (
                jsonify(
                    {"error": "Missing required fields: firstName, lastName, email"}
                ),
                400,
            )
        email_key = email_raw.casefold()
        phone = body.get("phone")
        phone_str = "" if phone is None else str(phone).strip()
        new_id = f"cust-{uuid.uuid4().hex[:8]}"
        now = _utc_now_iso()

        with _file_lock:
            rows = read_rows()
            for row in rows:
                if row["email"].strip().casefold() == email_key:
                    logger.info("POST duplicate email=%s -> 409", email_raw)
                    return jsonify({"error": "Email already exists"}), 409
            rows.append(
                {
                    "customer_id": new_id,
                    "first_name": str(first).strip(),
                    "last_name": str(last).strip(),
                    "email": email_raw,
                    "phone": phone_str,
                    "status": "ACTIVE",
                    "registered_at": now,
                }
            )
            write_rows(rows)
        logger.info("POST created customer_id=%s email=%s", new_id, email_raw)

        resp = jsonify(
            {"customerId": new_id, "message": "Customer created successfully"}
        )
        resp.status_code = 201
        resp.headers["Location"] = f"/customers/{new_id}"
        return resp

    return app


bootstrap_data()
app = create_app()

if __name__ == "__main__":
    host = os.environ.get("BIND_HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "5001"))
    app.run(host=host, port=port, debug=False)
